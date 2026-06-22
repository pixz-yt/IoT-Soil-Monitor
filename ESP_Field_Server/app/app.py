from flask import Flask, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time

app = Flask("app")

DATA_FOLDER = "FData"
EXCEL_FILE = os.path.join(DATA_FOLDER, "readings.xlsx")

os.makedirs(DATA_FOLDER, exist_ok=True)

temp = 0
humidity = 0

temp_buffer = []
hum_buffer = []

last_update = time.time()

if not os.path.exists(EXCEL_FILE):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "Time",
        "Average Temperature",
        "Average Humidity"
    ])

    wb.save(EXCEL_FILE)


@app.route("/")
def home():
    return send_file("dashboard.html")


@app.route("/logo.png")
def logo():
    return send_file("logo.png")


@app.route("/data", methods=["POST"])
def data():

    global temp
    global humidity
    global last_update

    d = request.json

    temp = float(d["temp"])
    humidity = float(d["humidity"])

    last_update = time.time()

    temp_buffer.append(temp)
    hum_buffer.append(humidity)

    if len(temp_buffer) >= 15:

        avg_temp = sum(temp_buffer) / len(temp_buffer)
        avg_hum = sum(hum_buffer) / len(hum_buffer)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            round(avg_temp, 2),
            round(avg_hum, 2)
        ])

        wb.save(EXCEL_FILE)

        temp_buffer.clear()
        hum_buffer.clear()

    return jsonify({
        "status": "ok"
    })


@app.route("/live")
def live():

    sensor_status = "ONLINE"

    if time.time() - last_update > 15:
        sensor_status = "OFFLINE"

    return jsonify({
        "temp": temp,
        "humidity": humidity,
        "samples": len(temp_buffer),
        "status": sensor_status
    })


app.run(
    host="0.0.0.0",
    port=5000,
    debug=True
)